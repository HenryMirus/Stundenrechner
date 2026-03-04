"""
Excel-Export-Modul für den Stundenrechner.
Erstellt professionell formatierte XLSX-Dateien mit Stundenübersichten.
"""

from collections import OrderedDict
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

GERMAN_MONTHS = [
    "Januar", "Februar", "März", "April", "Mai", "Juni",
    "Juli", "August", "September", "Oktober", "November", "Dezember",
]


class ExcelExporter:
    """Exportiert Stundeneinträge als formatierte Excel-Datei."""

    TITLE_FONT = Font(bold=True, size=14, name="Calibri", color="2C3E50")
    HEADER_FONT = Font(bold=True, size=10, name="Calibri", color="FFFFFF")
    HEADER_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    DATE_FONT = Font(bold=True, size=10, name="Calibri", color="2C3E50")
    NORMAL_FONT = Font(size=10, name="Calibri")
    DAY_TOTAL_FONT = Font(bold=True, size=10, name="Calibri", color="2C3E50")
    DAY_TOTAL_FILL = PatternFill(start_color="D5E8D4", end_color="D5E8D4", fill_type="solid")
    MONTH_TOTAL_FONT = Font(bold=True, size=11, name="Calibri", color="FFFFFF")
    MONTH_TOTAL_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    ALT_ROW_FILL = PatternFill(start_color="F7F9FC", end_color="F7F9FC", fill_type="solid")

    THIN_BORDER = Border(
        left=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )

    CENTER = Alignment(horizontal="center", vertical="center")
    LEFT = Alignment(horizontal="left", vertical="center")
    RIGHT = Alignment(horizontal="right", vertical="center")

    @classmethod
    def export(cls, entries: list, year: int, month: int, filepath: str):
        """
        Exportiert Monats-Einträge als XLSX-Datei.

        Args:
            entries: Liste von (id, date_iso, task, hours) Tupeln.
            year: Jahr.
            month: Monat (1-12).
            filepath: Pfad für die Ausgabedatei.
        """
        wb = Workbook()
        ws = wb.active
        month_name = GERMAN_MONTHS[month - 1]
        ws.title = f"{month_name} {year}"

        # Spaltenbreiten
        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 38
        ws.column_dimensions["E"].width = 14

        NUM_COLS = 5

        row = 1

        # ── Titel ────────────────────────────────────────────
        title_cell = ws.cell(
            row=row, column=1,
            value=f"Stundenübersicht – {month_name} {year}",
        )
        title_cell.font = cls.TITLE_FONT
        title_cell.alignment = cls.LEFT
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NUM_COLS)
        row += 2

        # ── Spaltenüberschriften ──────────────────────────────
        headers = ["Datum", "Kunde", "Komissions-Nr.", "Aufgabe", "Stunden"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = cls.HEADER_FONT
            cell.fill = cls.HEADER_FILL
            cell.alignment = cls.CENTER
            cell.border = cls.THIN_BORDER
        row += 1

        # ── Einträge nach Datum gruppieren ────────────────────
        entries_by_date = OrderedDict()
        for _id, date_str, task, hours, customer, commission in entries:
            entries_by_date.setdefault(date_str, []).append((customer, commission, task, hours))

        monthly_total = 0.0
        date_count = 0
        weekday_names = ["Mo", "Di", "Mi", "Do", "Fr", "Sa", "So"]

        for date_str, day_entries in entries_by_date.items():
            try:
                dt = datetime.strptime(date_str, "%Y-%m-%d")
                display_date = f"{weekday_names[dt.weekday()]}, {dt.strftime('%d.%m.%Y')}"
            except ValueError:
                display_date = date_str

            day_total = 0.0
            is_first = True
            use_alt = date_count % 2 == 1

            for customer, commission, task, hours in day_entries:
                cell_date = ws.cell(
                    row=row, column=1,
                    value=display_date if is_first else "",
                )
                cell_date.font = cls.DATE_FONT if is_first else cls.NORMAL_FONT
                cell_date.alignment = cls.LEFT
                cell_date.border = cls.THIN_BORDER
                if use_alt:
                    cell_date.fill = cls.ALT_ROW_FILL

                cell_customer = ws.cell(row=row, column=2, value=customer)
                cell_customer.font = cls.NORMAL_FONT
                cell_customer.alignment = cls.LEFT
                cell_customer.border = cls.THIN_BORDER
                if use_alt:
                    cell_customer.fill = cls.ALT_ROW_FILL

                cell_commission = ws.cell(row=row, column=3, value=commission)
                cell_commission.font = cls.NORMAL_FONT
                cell_commission.alignment = cls.LEFT
                cell_commission.border = cls.THIN_BORDER
                if use_alt:
                    cell_commission.fill = cls.ALT_ROW_FILL

                cell_task = ws.cell(row=row, column=4, value=task)
                cell_task.font = cls.NORMAL_FONT
                cell_task.alignment = cls.LEFT
                cell_task.border = cls.THIN_BORDER
                if use_alt:
                    cell_task.fill = cls.ALT_ROW_FILL

                cell_hours = ws.cell(row=row, column=5, value=hours)
                cell_hours.font = cls.NORMAL_FONT
                cell_hours.number_format = "0.00"
                cell_hours.alignment = cls.CENTER
                cell_hours.border = cls.THIN_BORDER
                if use_alt:
                    cell_hours.fill = cls.ALT_ROW_FILL

                day_total += hours
                is_first = False
                row += 1

            # ── Tagesgesamt ──────────────────────────────────
            for col_idx in range(1, NUM_COLS + 1):
                cell = ws.cell(row=row, column=col_idx)
                cell.fill = cls.DAY_TOTAL_FILL
                cell.font = cls.DAY_TOTAL_FONT
                cell.border = cls.THIN_BORDER

            ws.cell(row=row, column=1).value = ""
            ws.cell(row=row, column=4).value = "Tagesgesamt"
            ws.cell(row=row, column=4).alignment = cls.RIGHT
            cell_day_total = ws.cell(row=row, column=5, value=day_total)
            cell_day_total.number_format = "0.00"
            cell_day_total.alignment = cls.CENTER
            row += 2  # Leerzeile zwischen Tagen

            monthly_total += day_total
            date_count += 1

        # ── Monatsgesamt ──────────────────────────────────────
        for col_idx in range(1, NUM_COLS + 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.fill = cls.MONTH_TOTAL_FILL
            cell.font = cls.MONTH_TOTAL_FONT
            cell.border = cls.THIN_BORDER

        ws.cell(row=row, column=1).value = ""
        ws.cell(row=row, column=4).value = "MONATSGESAMT"
        ws.cell(row=row, column=4).alignment = cls.RIGHT
        cell_month_total = ws.cell(row=row, column=5, value=monthly_total)
        cell_month_total.number_format = "0.00"
        cell_month_total.alignment = cls.CENTER

        ws.page_margins.left = 0.5
        ws.page_margins.right = 0.5
        ws.print_area = f"A1:E{row}"

        wb.save(filepath)
